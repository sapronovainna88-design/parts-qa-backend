# app.py — FastAPI backend for Railway (robust, aliases + safe preview/process)
# -*- coding: utf-8 -*-

import os
import uuid
import re
from typing import Optional
from urllib.request import urlopen, Request

from fastapi import FastAPI, UploadFile, File, Form, Body, Header
from fastapi.responses import JSONResponse, FileResponse
from fastapi.middleware.cors import CORSMiddleware

import pandas as pd
from difflib import get_close_matches, SequenceMatcher
from tabulate import tabulate
import json
import openpyxl
from openpyxl.styles import PatternFill

# ----------------------------- Config -----------------------------
APP_PORT = int(os.getenv("PORT", "8000"))
UNIFICATION_PATH = os.getenv("UNIFICATION_PATH", "unifikatsiya.xlsx")
TMP_DIR = os.getenv("TMP_DIR", "tmp")
API_KEY = os.getenv("X_API_KEY")  # optional, header: X-API-Key
os.makedirs(TMP_DIR, exist_ok=True)

app = FastAPI(title="Parts QA Backend", version="1.2")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ----------------------- Text utils & brand aliases -----------------
BRAND_ALIASES = {
    "john deere": {"jd", "john-deere", "john  deere", "джон дир", "джон дір", "джондір", "джон-дир"},
    "horsch": {"horsh", "хорш", "horschmaschinen", "horsch-maschinen"},
    "claas": {"клас", "класс"},
    "case ih": {"case", "кейс", "кейс ih", "case-ih"},
    "new holland": {"нью холланд", "нью-холланд", "nh", "нх"},
    "vaderstad": {"вадерштад", "вадерстад", "vader-stad"},
}

TYPE_ALIASES = {
    "сеялки": "посівна техніка",
    "сеялка": "посівна техніка",
    "посевная техника": "посівна техніка",
}


def _strip(s: str) -> str:
    return str(s or "").strip()


def _norm_text(s: str) -> str:
    return re.sub(r"\s+", " ", _strip(s).lower())


def _brand_key(s: str) -> str:
    k = _norm_text(s)
    k = k.replace("-", " ")
    k = re.sub(r"[^a-z0-9\u0400-\u04FF ]+", "", k)
    k = re.sub(r"\s+", " ", k).strip()
    for canon, aliases in BRAND_ALIASES.items():
        if k == canon or k in aliases:
            return canon
    if k == "jd":
        return "john deere"
    return k


def _same_brand(a: str, b: str) -> bool:
    return _brand_key(a) == _brand_key(b)


def _pn_variants(pn: str):
    pn = _strip(pn)
    if not pn:
        return []
    v = {pn, pn.replace(".", ""), pn.replace("-", ""), re.sub(r"[.\-\s]", "", pn)}
    return list(v)


def _has_cyr(s: str) -> bool:
    return bool(re.search("[\u0400-\u04FF]", str(s)))


# ----------------------- Unification loading -----------------
# Optional synonym mapping (in case headers differ slightly)
_COLUMN_CANDIDATES = {
    "Вид техніки": ["вид техніки", "вид техники", "тип техніки", "тип техники", "категорія", "категория", "вид"],
    "Бренд": ["бренд", "виробник", "производитель", "виробник/бренд", "brand"],
    "Найменування": ["найменування", "наименование", "назва", "название", "name"],
    "каталожний номер": ["каталожний номер", "каталожный номер", "артикул", "part number", "pn", "код"],
    "Допустимі аналоги": ["допустимі аналоги", "допустимые аналоги", "аналоги", "equivalents", "alternative"],
}


def _find_col(df: pd.DataFrame, canonical: str) -> str:
    if canonical in df.columns:
        return canonical
    want = [_norm_text(x) for x in [canonical] + _COLUMN_CANDIDATES.get(canonical, [])]
    colmap = {orig: _norm_text(orig) for orig in df.columns}
    # exact
    for w in want:
        for orig, normed in colmap.items():
            if normed == w:
                return orig
    # contains
    for w in want:
        for orig, normed in colmap.items():
            if w in normed:
                return orig
    raise KeyError(f"Колонку для «{canonical}» не знайдено. Є: {list(df.columns)}")


def load_unification() -> pd.DataFrame:
    df = None
    last_err = None
    for p in [UNIFICATION_PATH, os.path.join(os.getcwd(), UNIFICATION_PATH), f"/app/{UNIFICATION_PATH}"]:
        try:
            df = pd.read_excel(p, header=0)
            break
        except Exception as e:
            last_err = e
            df = None
    if df is None:
        raise RuntimeError(f"Не вдалося прочитати уніфікацію з '{UNIFICATION_PATH}': {last_err}")

    # trim headers and map to canonical
    df.columns = [str(c).strip() for c in df.columns]
    # try to map major columns to canonical names
    try:
        cat_col = _find_col(df, "Вид техніки")
        brand_col = _find_col(df, "Бренд")
        name_col = _find_col(df, "Найменування")
        pn_col = _find_col(df, "каталожний номер")
        analogs_col = _find_col(df, "Допустимі аналоги")
    except Exception:
        # fallback: assume exact ukr headers
        cat_col, brand_col, name_col, pn_col, analogs_col = (
            "Вид техніки", "Бренд", "Найменування", "каталожний номер", "Допустимі аналоги"
        )

    df = df.rename(
        columns={
            cat_col: "Вид техніки",
            brand_col: "Бренд",
            name_col: "Найменування",
            pn_col: "каталожний номер",
            analogs_col: "Допустимі аналоги",
        }
    )

    # normalized fields
    df["cat_norm"] = df["Вид техніки"].astype(str).str.strip().str.lower()
    df["brand_norm"] = df["Бренд"].astype(str).str.strip().str.lower()
    df["brand_key"] = df["brand_norm"].apply(_brand_key)
    return df


def _require_api_key(x_api_key: Optional[str]):
    if API_KEY and x_api_key != API_KEY:
        raise PermissionError("unauthorized")


# ===== META: lists for GPT UI =================================================
@app.get("/meta/types")
def meta_types(x_api_key: Optional[str] = Header(default=None, convert_underscores=False)):
    try:
        _require_api_key(x_api_key)
    except PermissionError:
        return JSONResponse({"error": "unauthorized"}, status_code=401)

    try:
        df = load_unification()
    except Exception:
        return {"items": []}

    out = (
        df[["cat_norm", "Вид техніки"]]
        .dropna()
        .drop_duplicates(subset=["cat_norm"])  # unique by normalized key
        .sort_values("Вид техніки")
        ["Вид техніки"]
        .tolist()
    )
    return {"items": out}


@app.get("/meta/brands")
def meta_brands(
    type: Optional[str] = None,
    x_api_key: Optional[str] = Header(default=None, convert_underscores=False),
):
    try:
        _require_api_key(x_api_key)
    except PermissionError:
        return JSONResponse({"error": "unauthorized"}, status_code=401)

    try:
        df = load_unification()
    except Exception:
        return {"items": []}

    if type:
        st = _norm_text(TYPE_ALIASES.get(_norm_text(type), type))
        df = df[df["cat_norm"] == st]

    items = sorted(set(df["Бренд"].dropna().astype(str).str.strip().tolist()))
    if "Всі доступні моделі" not in items:
        items = ["Всі доступні моделі"] + items
    return {"items": items}


# ------------------------------- Health ---------------------------------
@app.get("/healthz")
def healthz():
    return {"ok": True}


# -------------- Robust brand picking (exact -> strict fuzzy -> all) --------------

def _pick_brand(df_cat: pd.DataFrame, selected_brand_raw: str):
    """Return (used_brand_key, brand_mode, normalized_brand_label)."""
    sb_raw = _strip(selected_brand_raw)
    sb_key = _brand_key(sb_raw)
    norm_all = _norm_text("всі доступні моделі")

    if sb_key in {"всі", "все", "all", norm_all}:
        return None, "all_manual", "Всі доступні моделі"

    pool_keys = df_cat["brand_key"].dropna().unique().tolist()

    # exact
    if sb_key in pool_keys:
        human = df_cat.loc[df_cat["brand_key"] == sb_key, "Бренд"].iloc[0]
        return sb_key, "exact", str(human)

    # strict fuzzy (0.85)
    best_key, best_ratio = None, 0.0
    for k in pool_keys:
        r = SequenceMatcher(None, sb_key, k).ratio()
        if r > best_ratio:
            best_ratio, best_key = r, k
    if best_key and best_ratio >= 0.85:
        human = df_cat.loc[df_cat["brand_key"] == best_key, "Бренд"].iloc[0]
        return best_key, "fuzzy_exact", str(human)

    # fallback: 'всі'
    has_all = (_norm_text("всі доступні моделі") in df_cat["brand_norm"].values)
    if has_all:
        return None, "all_fallback_no_exact", "Всі доступні моделі"

    return None, "none_found_unification_skipped", (sb_raw or "—")


# ------------------------------- PREVIEW ---------------------------------
@app.post("/preview")
async def preview(
    selected_type: str = Form(...),
    selected_brand: str = Form(...),
    main_file: UploadFile = File(...),
    x_api_key: Optional[str] = Header(default=None, convert_underscores=False),
):
    try:
        _require_api_key(x_api_key)
    except PermissionError:
        return JSONResponse({"error": "unauthorized"}, status_code=401)

    token = str(uuid.uuid4())
    tmp_main = os.path.join(TMP_DIR, f"{token}.xlsx")
    content = await main_file.read()
    with open(tmp_main, "wb") as f:
        f.write(content)

    try:
        df = load_unification()
    except Exception as e:
        return JSONResponse({"error": f"Помилка уніфікації: {e}"}, status_code=400)

    st_raw = (selected_type or "").strip()
    sb_raw = (selected_brand or "").strip()

    # normalize type with aliases
    st_alias = TYPE_ALIASES.get(_norm_text(st_raw), _norm_text(st_raw))
    df_cat = df[df["cat_norm"] == st_alias]
    if df_cat.empty:
        choices = df["cat_norm"].dropna().unique().tolist()
        suggestions = get_close_matches(_norm_text(st_raw), choices, n=3, cutoff=0.6)
        return JSONResponse({"error": f"Категорію «{st_raw}» не знайдено", "suggestions": suggestions}, status_code=400)
    normalized_type = df_cat["Вид техніки"].iloc[0]

    # brand pick
    used_brand_key, brand_mode, normalized_brand = _pick_brand(df_cat, sb_raw)

    norm_all = _norm_text("всі доступні моделі")
    df_all = df_cat[df_cat["brand_norm"] == norm_all]
    if used_brand_key:
        df_exact = df_cat[df_cat["brand_key"] == used_brand_key]
        df_brand = pd.concat([df_exact, df_all], ignore_index=True) if not df_all.empty else df_exact
    else:
        df_brand = df_all if not df_all.empty else pd.DataFrame(
            columns=["Найменування", "каталожний номер", "Допустимі аналоги"]
        )

    # build preview temp table
    if df_brand.empty:
        temp_df = pd.DataFrame(columns=["C (Найменування)", "D (Каталожний номер)", "E (Допустимі аналоги)"])
    else:
        temp_df = (
            df_brand[["Найменування", "каталожний номер", "Допустимі аналоги"]]
            .drop_duplicates()
            .reset_index(drop=True)
            .rename(columns={
                "Найменування": "C (Найменування)",
                "каталожний номер": "D (Каталожний номер)",
                "Допустимі аналоги": "E (Допустимі аналоги)",
            })
        )

    tmp_temp = os.path.join(TMP_DIR, f"{token}_temp.parquet")
    temp_df.to_parquet(tmp_temp, index=False)

    # save meta for process (remember chosen brand & mode)
    try:
        tmp_meta = os.path.join(TMP_DIR, f"{token}_meta.json")
        with open(tmp_meta, "w", encoding="utf-8") as mf:
            json.dump({
                "normalized_type": str(normalized_type),
                "normalized_brand": str(normalized_brand),
                "brand_mode": str(brand_mode),
            }, mf, ensure_ascii=False)
    except Exception:
        pass

    if temp_df.empty:rocess (remember chosen brand & mode)
    try:
        tmp_meta = os.path.join(TMP_DIR, f"{token}_meta.json")
        with open(tmp_meta, "w", encoding="utf-8") as mf:
            json.dump({
                "normalized_type": str(normalized_type),
                "normalized_brand": str(normalized_brand),
                "brand_mode": str(brand_mode),
            }, mf, ensure_ascii=False)
    except Exception:
        pass

    if temp_df.empty:
        preview_markdown = "_Уніфікацію не знайдено — обробка піде без уніфікації._"
    else:
        preview_markdown = tabulate(temp_df, headers=temp_df.columns, tablefmt="github", showindex=False)

    return {
        "normalized_type": normalized_type,
        "normalized_brand": normalized_brand,
        "brand_mode": brand_mode,
        "preview_markdown": preview_markdown,
        "token": token,
    }


# ------------------------------- PREVIEW by URL (fallback) --------------------
@app.post("/preview_url")
async def preview_url(
    payload: dict = Body(...),
    x_api_key: Optional[str] = Header(default=None, convert_underscores=False),
):
    try:
        _require_api_key(x_api_key)
    except PermissionError:
        return JSONResponse({"error": "unauthorized"}, status_code=401)

    selected_type = str(payload.get("selected_type", "")).strip()
    selected_brand = str(payload.get("selected_brand", "")).strip()
    file_url = str(payload.get("file_url", "")).strip()
    if not selected_type or not selected_brand or not file_url:
        return JSONResponse({"error": "selected_type, selected_brand і file_url обов'язкові"}, status_code=400)

    token = str(uuid.uuid4())
    os.makedirs(TMP_DIR, exist_ok=True)
    tmp_main = os.path.join(TMP_DIR, f"{token}.xlsx")
    try:
        req = Request(file_url, headers={"User-Agent": "Mozilla/5.0"})
        with urlopen(req, timeout=60) as resp, open(tmp_main, "wb") as f:
            f.write(resp.read())
    except Exception as e:
        return JSONResponse({"error": f"Не вдалося завантажити файл за URL: {e}"}, status_code=400)

    # reuse preview logic
    try:
        df = load_unification()
    except Exception as e:
        return JSONResponse({"error": f"Помилка уніфікації: {e}"}, status_code=400)

    st_raw = selected_type
    sb_raw = selected_brand

    st_alias = TYPE_ALIASES.get(_norm_text(st_raw), _norm_text(st_raw))
    df_cat = df[df["cat_norm"] == st_alias]
    if df_cat.empty:
        choices = df["cat_norm"].dropna().unique().tolist()
        suggestions = get_close_matches(_norm_text(st_raw), choices, n=3, cutoff=0.6)
        return JSONResponse({"error": f"Категорію «{st_raw}» не знайдено", "suggestions": suggestions}, status_code=400)
    normalized_type = df_cat["Вид техніки"].iloc[0]

    used_brand_key, brand_mode, normalized_brand = _pick_brand(df_cat, sb_raw)

    norm_all = _norm_text("всі доступні моделі")
    df_all = df_cat[df_cat["brand_norm"] == norm_all]
    if used_brand_key:
        df_exact = df_cat[df_cat["brand_key"] == used_brand_key]
        df_brand = pd.concat([df_exact, df_all], ignore_index=True) if not df_all.empty else df_exact
    else:
        df_brand = df_all if not df_all.empty else pd.DataFrame(
            columns=["Найменування", "каталожний номер", "Допустимі аналоги"]
        )

    if df_brand.empty:
        temp_df = pd.DataFrame(columns=["C (Найменування)", "D (Каталожний номер)", "E (Допустимі аналоги)"])
    else:
        temp_df = (
            df_brand[["Найменування", "каталожний номер", "Допустимі аналоги"]]
            .drop_duplicates()
            .reset_index(drop=True)
            .rename(columns={
                "Найменування": "C (Найменування)",
                "каталожний номер": "D (Каталожний номер)",
                "Допустимі аналоги": "E (Допустимі аналоги)",
            })
        )

    tmp_temp = os.path.join(TMP_DIR, f"{token}_temp.parquet")
    temp_df.to_parquet(tmp_temp, index=False)

    if temp_df.empty:
        preview_markdown = "_Уніфікацію не знайдено — обробка піде без уніфікації._"
    else:
        preview_markdown = tabulate(temp_df, headers=temp_df.columns, tablefmt="github", showindex=False)

    return {
        "normalized_type": normalized_type,
        "normalized_brand": normalized_brand,
        "brand_mode": brand_mode,
        "preview_markdown": preview_markdown,
        "token": token,
    }


# ------------------------------- PROCESS ---------------------------------
@app.post("/process")
async def process(
    token: Optional[str] = Form(None),
    payload: Optional[dict] = Body(None),
    x_api_key: Optional[str] = Header(default=None, convert_underscores=False),
):
    try:
        _require_api_key(x_api_key)
    except PermissionError:
        return JSONResponse({"error": "unauthorized"}, status_code=401)

    if not token and isinstance(payload, dict):
        token = payload.get("token")
    if not token:
        return JSONResponse({"error": "token is required"}, status_code=400)

    tmp_main = os.path.join(TMP_DIR, f"{token}.xlsx")
    tmp_temp = os.path.join(TMP_DIR, f"{token}_temp.parquet")
    if not os.path.exists(tmp_main):
        return JSONResponse({"error": "invalid token"}, status_code=400)

    # read preview temp table
    temp_df = pd.DataFrame()
    if os.path.exists(tmp_temp):
        try:
            temp_df = pd.read_parquet(tmp_temp)
        except Exception:
            temp_df = pd.DataFrame()

    unification_enabled = not temp_df.empty

    wb = openpyxl.load_workbook(tmp_main)
    ws = wb.active

    red = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    yellow = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
    none = PatternFill()

    # Build header index by fuzzy contains
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    norm_headers = [_norm_text(h) for h in headers]

    def _find_idx(cands):
        for i, h in enumerate(norm_headers):
            for c in cands:
                if c in h:
                    return i
        return None

    idx_name = _find_idx(["наймен", "наимен", "назва", "название", "name"]) or 0
    idx_desc = _find_idx(["опис", "описание", "description"]) or (1 if len(headers) > 1 else 0)
    idx_brand = _find_idx(["виробник/бренд", "виробник", "производитель", "бренд", "brand"]) or (2 if len(headers) > 2 else 0)
    idx_pn_sup = _find_idx(["каталожний номер виробника", "артикул виробника", "каталожний номер", "артикул", "part number", "pn"]) or (3 if len(headers) > 3 else 0)
    idx_orig = _find_idx(["оригінал/аналог", "оригинал/аналог"]) or (4 if len(headers) > 4 else 0)

    temp_records = temp_df.to_dict("records")

    # We use normalized_brand from preview to check 'original' lines
    # If temp table absent, fallback to empty string (won't enforce brand on originals)
    normalized_brand = ""
    # read normalized_brand saved at preview stage
    tmp_meta = os.path.join(TMP_DIR, f"{token}_meta.json")
    if os.path.exists(tmp_meta):
        try:
            with open(tmp_meta, "r", encoding="utf-8") as mf:
                m = json.load(mf)
                normalized_brand = str(m.get("normalized_brand", ""))
        except Exception:
            normalized_brand = ""
        except Exception:
            pass

    red_rows = yellow_rows = 0

    for row in ws.iter_rows(min_row=2):
        # reset fills
        for cell in row:
            cell.fill = none

        A = _strip(row[idx_name].value)
        D = _strip(row[idx_desc].value)
        J = _strip(row[idx_brand].value)
        L = _strip(row[idx_pn_sup].value)
        P = _strip(row[idx_orig].value).lower()

        had_red = False
        had_yellow = False

        if P == "аналог":
            if not unification_enabled:
                pass  # no highlighting when unification missing
            else:
                matched = False
                allowed: list[str] = []

                # 1) PN in description or equals to supplier PN (with variants)
                for r in temp_records:
                    pn = _strip(r.get("D (Каталожний номер)", ""))
                    if not pn:
                        continue
                    variants = _pn_variants(pn)
                    if any(v and v in D for v in variants) or (L and (L in variants or re.sub(r"[.\-\s]", "", L) in variants)):
                        matched = True
                        E = _strip(r.get("E (Допустимі аналоги)", ""))
                        allowed = [x.strip() for x in E.split(",") if x.strip()]
                        break

                ratio = 0.0
                if not matched and len(temp_records) > 0:
                    base = A.split(",")[0].split("|")[0].strip().lower()
                    best_idx = -1
                    best_ratio = -1
                    for i, r in enumerate(temp_records):
                        cval = str(r.get("C (Найменування)", "")).lower()
                        rratio = SequenceMatcher(None, base, cval).ratio()
                        if rratio > best_ratio:
                            best_ratio = rratio
                            best_idx = i
                    ratio = best_ratio if best_idx >= 0 else 0.0
                    if best_idx >= 0:
                        E = _strip(temp_records[best_idx].get("E (Допустимі аналоги)", ""))
                        allowed = [x.strip() for x in E.split(",") if x.strip()]
                    if ratio < 0.6:
                        row[idx_orig].fill = red
                        had_red = True
                    elif ratio < 0.8:
                        row[idx_name].fill = yellow
                        row[idx_orig].fill = yellow
                        had_yellow = True

                # 3) Brand J vs allowed analogs
                if allowed and J:
                    ok = any(_same_brand(J, a) or SequenceMatcher(None, _norm_text(J), _norm_text(a)).ratio() >= 0.8 for a in allowed)
                    if not ok:
                        close = any(SequenceMatcher(None, _norm_text(J), _norm_text(a)).ratio() >= 0.6 for a in allowed)
                        row[idx_brand].fill = yellow if close else red
                        row[idx_orig].fill = yellow if close else red
                        had_yellow = had_yellow or close
                        had_red = had_red or (not close)
        else:
            # ORIGINAL line checks: brand & supplier PN in description
            # Brand: use aliases first, then similarity
            if normalized_brand:
                ref_brand = normalized_brand
            else:
                ref_brand = J  # if we don't know, don't enforce
            if not _same_brand(J, ref_brand):
                ratio_j = SequenceMatcher(None, _norm_text(J), _norm_text(ref_brand)).ratio()
                if ratio_j < 0.6:
                    row[idx_brand].fill = red
                    had_red = True
                elif ratio_j < 0.8 or _has_cyr(J):
                    row[idx_brand].fill = yellow
                    had_yellow = True
            if L and L not in D:
                row[idx_pn_sup].fill = red
                had_red = True

        if had_red:
            row[idx_name].fill = red
            red_rows += 1
        elif had_yellow:
            # keep yellow on name only if no red
            if row[idx_name].fill != red:
                row[idx_name].fill = yellow
            yellow_rows += 1

    out_name = f"{token}_processed.xlsx"
    out_path = os.path.join(TMP_DIR, out_name)
    wb.save(out_path)

    return {"download_url": f"/download/{token}"}


@app.get("/download/{token}")
async def download(token: str):
    out_name = f"{token}_processed.xlsx"
    path = os.path.join(TMP_DIR, out_name)
    if not os.path.exists(path):
        # fallback: find by prefix
        for fname in os.listdir(TMP_DIR):
            if fname.startswith(token) and fname.endswith("_processed.xlsx"):
                path = os.path.join(TMP_DIR, fname)
                break
    if not os.path.exists(path):
        return JSONResponse({"error": "not found"}, status_code=404)
    return FileResponse(
        path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=os.path.basename(path),
    )


# ------------------------------- Main (local run) ----------------------------
if __name__ == "__main__":
    import uvicorn

    uvicorn.run(app, host="0.0.0.0", port=APP_PORT)
